import re
import urllib.request
import configparser
from os import path

import openpyxl
import datetime
from bs4 import BeautifulSoup
import lxml
# cf = configparser.ConfigParser();
# cf.read("D:/test.ini");

# cf.add_section("task");
# cf.set("task","key1","value1");
# cf.write(open("D:/test.ini",'w'));
# se = cf.sections();
# op = cf.options(se[0]);
# its = cf.items(se[0]);
# value = cf.get(se[0],op[0]);
# print("key1" in cf.options("task"));

# try:
#     wb = openpyxl.load_workbook("FundInfo.xlsx");
#     ws = wb["Sheet"];
#     print(datetime.datetime.now().strftime("%Y"));
#     ws["A1"] = datetime.datetime.now().strftime("%F");
#     wb.save("FundInfo.xlsx");
# except Exception as e:
#     print(e);
#     wb = openpyxl.Workbook();
#     ws = wb.active;
#     ws.append([1,2,3]);
#     ws.append([4,5,6]);
#     wb.save("FundInfo.xlsx");

# time = datetime.datetime.now();
# t = time.strftime("%F");
# print(time.strftime("%S"))
# print(t);

# fundList = ['184801','000011','070002','260104'];
# for id in fundList:
#     response = urllib.request.urlopen("http://fund.eastmoney.com/" + id + ".html?spm=aladin");
#     html = response.read().decode("utf-8");
#     soup = BeautifulSoup(html,'lxml');
#     print(soup.title.string);
    # pattern = re.compile(r'\d.+\d');
    # print((re.search(pattern,soup.find_all(class_="fix_date")[0].string).group()),end = "");
    # print((soup.find_all(class_="fix_date")[0]).next_sibling.string);

# pattern = re.compile(r'\d.+\d');
# result = re.search(pattern,"(09-30):");
# print(result.group());

# print(path.exists("fund.ini"));

# print('184801,000011,070002,260104'.split(","));

# list = [];i
# for i in range(10):
#     if len(list) == 0:
#         print("haha");
#     list.append(i);
# print(list);

wb = openpyxl.Workbook();
ws = wb["Sheet"];
strs = "时间 and 净值";
ws["A1"] = strs;
ws.column_dimensions['A'].width = len(strs)*1.5;
ws["A2"] = 1;
ws["A3"] = 2;
ws["A4"] = 3;
ws["A5"] = '=AVERAGE(A2:A4)';
print(ws.max_row);
print(ws.max_column);
wb.save("hai.xlsx");


