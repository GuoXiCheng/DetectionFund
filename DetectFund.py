"""
任务目标：爬取每日基金净值，保存到本地Excel中
"""
import datetime
import re
import urllib.request
from os import path

import configparser, lxml

import openpyxl
from bs4 import BeautifulSoup

# 配置文件类
from openpyxl.utils import get_column_letter


class Config:
    updateTime = None;
    cf = configparser.ConfigParser();

    def __init__(self):
        return;

    # 创建一个配置文件
    def newConfig(self):
        self.cf.add_section("fund");
        self.cf.set("fund", "id", "");
        self.cf.add_section("date");
        self.cf.set("date", "time", "");
        self.cf.write(open("fundConfig.ini", 'w'));
        return

    # 读取配置文件中的时间，判断爬取到的基金的时间与配置文件中的时间是否一致
    # 一致就退出程序，不一致就继续执行
    def readTime(self):
        self.cf.read("fundConfig.ini");
        time = self.cf.get("date", "time");
        if time == self.updateTime:
            exit(0);
        else:
            return False;
        return

    # 程序运行结束前，将更新时间写入配置文件
    def writeTime(self):
        self.cf.set("date", "time", self.updateTime);
        self.cf.write(open("fundConfig.ini", 'w'));
        return

    # 读取配置文件中所有需要爬取的基金id，返回列表
    def readFundList(self):
        self.cf.read("fundConfig.ini");
        idStr = self.cf.get("fund", "id");
        idList = idStr.split(",");
        return idList;


# 爬取功能类
class CrawlUtil:
    idList = None;
    updateTime = None;

    def __init__(self, idList):
        self.idList = idList;
        return

    # 依据基金列表爬取数据，返回(基金名称,净值数据)元组组成的列表
    def getFundInfo(self):
        infoList = [];
        temp = [];
        pattern = re.compile(r'\d.+\d');
        for fundId in self.idList:
            response = urllib.request.urlopen("http://fund.eastmoney.com/" + fundId + ".html?spm=aladin");
            html = response.read().decode("utf-8");
            soup = BeautifulSoup(html, 'lxml');
            temp.append(soup.title.string.split(")")[0] + ")");  # 基金名
            node = soup.find_all(class_="fix_date")[0];
            if len(infoList) == 0:
                self.updateTime = datetime.datetime.now().strftime("%Y") + "-" + re.search(pattern,
                                                                                           node.string).group();
            temp.append(node.next_sibling.string);  # 净值数据
            infoList.append(tuple(temp));
            temp.clear();
        return infoList;


# Excel对象类
class ExcelObject:
    updateTime = None;
    infoList = None;
    wb = None;
    ws = None;
    titleList = None;

    def __init__(self, updateTime, infoList):
        self.updateTime = updateTime;
        self.infoList = infoList;
        return

    # 判断是否存在一个Excel，如果有就加载，没有就新建
    def existExcel(self):
        if path.exists("fundExcel.xlsx"):
            self.wb = openpyxl.load_workbook("fundExcel.xlsx");
            self.ws = self.wb["Sheet"];
        else:
            self.wb = openpyxl.Workbook();
            self.ws = self.wb.active;
        return

    #删除平均值那一行
    def delAvg(self):
        if self.ws["A" + str(self.ws.max_row)].value == "平均值":
            self.ws.delete_rows(self.ws.max_row);
        return

    # 向Excel中插入数据
    def insertValue(self):
        titleList = ["时间 and 净值"];
        dataList = [self.updateTime];
        for data in self.infoList:
            titleList.append(data[0]);
            dataList.append(float(data[1]));
        self.titleList = titleList;
        for index in range(len(titleList)):
            self.ws[get_column_letter(index + 1) + "1"] = titleList[index];
        self.ws.append(dataList);
        return

    # 向Excel中插入平均值公式
    def insertAvgForm(self):
        avgList = ["平均值"];
        for index in range(len(self.titleList) - 1):
            col = get_column_letter(index + 2);
            avgList.append("=AVERAGE(" + col + "2" + ":" + col + str(self.ws.max_row) + ")");
        self.ws.append(avgList);
        return

    # 依据某一行的字符个数去调整数据表的列宽
    def adjustColWidth(self):
        for index in range(len(self.titleList)):
            self.ws.column_dimensions[get_column_letter(index + 1)].width = len(self.titleList[index]) * 1.5;
        return

    # 保存Excel
    def saveExcel(self):
        self.wb.save("fundExcel.xlsx");
        return


# 读取配置文件基金id
cf = Config();
idList = cf.readFundList();

# 获取基金信息
cu = CrawlUtil(idList);
infoList = cu.getFundInfo();

# 判断基金是否更新
updateTime = cu.updateTime;
cf.updateTime = updateTime;

if not cf.readTime():
    eo = ExcelObject(updateTime,infoList);
    eo.existExcel();
    eo.delAvg();
    eo.insertValue();
    eo.insertAvgForm();
    eo.adjustColWidth();
    eo.saveExcel();
    cf.writeTime();